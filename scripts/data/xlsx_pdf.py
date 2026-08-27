#!/usr/bin/env python3
"""xlsx_pdf.py — XLSX → 可交付 PDF（打印排版件）

Why 立此 engine（2026-08-27）：总部原有 xlsx 能力（`data.py` 的 convert/merge/split/lower、
`data.py view/fig/to-db`、report 的 `df_to_xlsx_report`）全是「进 xlsx」或「读 xlsx」，
**没有一条「出交付 PDF」的路**。而直接拿 Excel 导出必踩三个坑（2026-08-27 花溪水质汇总表实测）：
① 列宽不够 → 表头被截成「取水量（≥」这种残字；② 列数多 → 月份列溢出到下一页；
③ 多 sheet 导出后每页看不出是哪张表。本 engine 把「打印前该做的事」固化成默认值。

默认行为（非交互，一条命令出件）：
  1. **排版预处理**（做在临时副本上，源 xlsx 一个字节不动）
     · CJK 感知列宽（中文按 2 字宽，上限 44）· 表头加粗 + 冻结首行 + 每页重复表头
     · 细边框 · 首列左对齐其余居中 · 单列的说明页改宽列 + 自动换行
     · 页眉 `&A`（sheet 名，多 sheet 时认得出哪页是谁）· 页脚「第 P 页 / 共 N 页」
     · A4 · fitToWidth=1 / fitToHeight=0（压到一页宽，纵向自然分页）
     · 方向 auto：列数 ≥ 6 横向，否则纵向
  2. **导出**：`document/sub/excel_render.py`（Microsoft Excel 真版面 + 页数门）。
     全局禁令：版面只认 Word / WPS / Excel，第三方渲染器一律不用。

fail-closed（铁律 #2）：无 sheet / 全空 → exit 2（拒绝在空集上报绿）；
--sheets 指名了不存在的 sheet → exit 2；导出后 PDF 不存在或 0 页 → exit 1。

用法:
  python3 xlsx_pdf.py 表.xlsx                       # → 表.pdf（同目录）
  python3 xlsx_pdf.py 表.xlsx --out /path/out.pdf
  python3 xlsx_pdf.py 表.xlsx --title "XX 汇总表"    # 页眉标题（与 sheet 名并排）
  python3 xlsx_pdf.py 表.xlsx --orientation portrait --sheets 花溪水库,马蹄坑水库
  python3 xlsx_pdf.py 表.xlsx --keep-xlsx 打印版.xlsx  # 顺带留一份排过版的 xlsx
  python3 xlsx_pdf.py 表.xlsx --no-fmt               # 原样导出，不动排版

统一 CLI 入口: python3 ~/Dev/tools/doctools/scripts/data/data.py xlsx-pdf 表.xlsx
"""
from __future__ import annotations

import argparse
import shutil
import sys
import tempfile
from pathlib import Path

_SELF = Path(__file__).resolve()
sys.path.insert(0, str(_SELF.parent.parent / "document" / "sub"))
from excel_render import render as excel_render  # noqa: E402

SCRIPT_VERSION = "1.0.0"

MAX_COL_WIDTH = 44
MIN_COL_WIDTH = 8
TEXT_COL_WIDTH = 100
LANDSCAPE_MIN_COLS = 6


def _cjk_width(s: str) -> int:
    return sum(2 if ord(ch) > 127 else 1 for ch in s)


def format_for_print(src: Path, dst: Path, title: str | None,
                     orientation: str, sheets: list[str] | None) -> int:
    """在副本上做打印排版；返回排版过的 sheet 数（-1 = --sheets 指名不存在）。源文件不动。"""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.worksheet.properties import PageSetupProperties

    wb = openpyxl.load_workbook(src)
    if sheets:
        unknown = set(sheets) - set(wb.sheetnames)
        if unknown:
            print(f"[xlsx_pdf] --sheets 指名了不存在的 sheet: {sorted(unknown)}；"
                  f"本表实有 {wb.sheetnames}", file=sys.stderr)
            return -1
        for name in list(wb.sheetnames):
            if name not in set(sheets):
                del wb[name]

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    done = 0

    for ws in wb.worksheets:
        ncols, nrows = ws.max_column or 0, ws.max_row or 0
        # ⚠ 空 sheet 的 max_row/max_column 也是 1，不是 0 —— 只看维度会把空表当有料，
        # 导出一页白纸然后 exit 0（2026-08-27 反向验证抓到）。必须看有没有真值。
        has_data = any(c.value is not None for row in ws.iter_rows() for c in row)
        if ncols == 0 or nrows == 0 or not has_data:
            continue
        text_only = ncols == 1          # 说明页这类：宽列 + 换行，不加表格线

        for col in ws.iter_cols():
            width = max((_cjk_width(str(c.value)) for c in col if c.value is not None),
                        default=0)
            ws.column_dimensions[col[0].column_letter].width = (
                TEXT_COL_WIDTH if text_only
                else min(max(width + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
            )

        for row in ws.iter_rows():
            for cell in row:
                if text_only:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="left" if cell.column == 1 else "center",
                        vertical="center")
        if not text_only:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "B2"
            ws.print_title_rows = "1:1"

        landscape = (orientation == "landscape") or (
            orientation == "auto" and ncols >= LANDSCAPE_MIN_COLS)
        ws.page_setup.orientation = "landscape" if landscape else "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.oddHeader.center.text = f"{title}　&A" if title else "&A"
        ws.oddHeader.center.size = 9
        ws.oddFooter.right.text = "第 &P 页 / 共 &N 页"
        ws.oddFooter.right.size = 9
        done += 1

    if done:
        wb.save(dst)
    return done


def run(args: argparse.Namespace) -> int:
    src = Path(args.xlsx).expanduser().resolve()
    if not src.is_file():
        print(f"[xlsx_pdf] 源文件不存在: {src}", file=sys.stderr)
        return 2
    out = Path(args.out).expanduser().resolve() if args.out else src.with_suffix(".pdf")
    sheets = [s.strip() for s in args.sheets.split(",") if s.strip()] if args.sheets else None

    with tempfile.TemporaryDirectory() as td:
        staged = Path(td) / src.name
        if args.no_fmt:
            shutil.copy2(src, staged)
        else:
            n = format_for_print(src, staged, args.title, args.orientation, sheets)
            if n < 0:
                return 2
            if n == 0:
                print("[xlsx_pdf] 没有一个非空 sheet，拒绝出件", file=sys.stderr)
                return 2
            print(f"[xlsx_pdf] 排版 {n} 个 sheet（方向 {args.orientation}）")
        if args.keep_xlsx:
            keep = Path(args.keep_xlsx).expanduser().resolve()
            keep.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(staged, keep)
            print(f"[xlsx_pdf] 排版件留档 {keep}")
        rc = excel_render(staged, out, min_pages=args.min_pages)
    return rc


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="xlsx_pdf",
        description="XLSX → 可交付 PDF（打印排版件；Excel 真版面）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="用法:" + __doc__.split("用法:")[-1])
    p.add_argument("xlsx", help="源 .xlsx（只读，不会被改）")
    p.add_argument("--out", help="输出 PDF 路径（默认同目录同名 .pdf）")
    p.add_argument("--title", help="页眉标题（与 sheet 名并排；默认只出 sheet 名）")
    p.add_argument("--orientation", choices=["auto", "portrait", "landscape"], default="auto",
                   help=f"页面方向（auto = 列数 ≥ {LANDSCAPE_MIN_COLS} 走横向）")
    p.add_argument("--sheets", help="只导出这些 sheet（逗号分隔）")
    p.add_argument("--keep-xlsx", help="把排过版的 xlsx 也留一份到此路径")
    p.add_argument("--no-fmt", action="store_true", help="不做排版预处理，原样导出")
    p.add_argument("--min-pages", type=int, default=1, help="产物页数门（低于即判失败）")
    p.add_argument("--version", action="version", version=f"xlsx_pdf {SCRIPT_VERSION}")
    return p


def main(argv: list[str] | None = None) -> int:
    return run(build_parser().parse_args(argv))


if __name__ == "__main__":
    sys.exit(main())
